#encoding: UTF-8
import openpyxl
import requests
import xlrd
import pandas as pd
import config
import imaplib
import os
from invoice import Invoice
import json

def def_connect(section_box): # INBOX или Spam
    imap_server = config.imap_server
    imap = imaplib.IMAP4_SSL(imap_server)
    try:
        connect = imap.login(config.provider_mail_login, config.provider_mail_password)
        for i in connect:
            if i == 'OK':
                print(f'{i}, Доступ к почте получен!\n')
            break
    except:
            print(f"Не удалось подключиться к почтовому ящику {config.provider_mail_login}")

    imap.select(section_box)
    return imap

def last_uid_list(def_connect, amount, letter_status):
    """Возвращает список последних amount(int) писем со статусом прочтения letter_status ("ALL" или "SEEN"). def_connect: объект imap"""
    mess_UID = def_connect.uid('search', letter_status, "ALL")
    string_UID = mess_UID[1][0].decode()
    listUID = [num.encode() for num in string_UID.split()]
    last_uid_list = listUID[:(len(listUID) - amount):-1]
    return last_uid_list

def delete_all_files(dir):
    all_files = os.listdir(dir)
    for file_name in all_files:
        os.remove(f'{dir}\\{file_name}')

def invoices_lists(dictpath=config.inbox_dict_path):
    """Функция возвращает список всех накладных и xlsx файлов в папке dictpath"""
    all_invoices_list = []
    xlsx_files = []
    all_files = os.listdir(dictpath)
    for file_name in all_files:
        file_name = config.inbox_dict_path + '\\' + file_name
        all_invoices_list.append(file_name)
        if file_name[-1] == 'x':
            xlsx_files.append(file_name)
    return all_invoices_list, xlsx_files

def check_invoice(invalid_list_buyer=config.invalid_buyer):
    """Проверяет наличие покупателя в списке. Если покупатель в списке - НЕ выгружает накладную и удаляет ее"""
    all_invoices_list, xlsx_files = invoices_lists()
    for file in xlsx_files:
        Invoice_object = Invoice(file)
        invoice_buyer = Invoice_object.buyer
        for inv_buyer in invalid_list_buyer:
            if inv_buyer.upper() in invoice_buyer.upper():
                os.remove(file)

def refactor_date_data(date):
    date_dict = {
        "Января": '01',
        "Февраля": '02',
        "Марта": '03',
        "Апреля": '04',
        "Мая": '05',
        "Июня": '06',
        "Июля": '07',
        "Августа": '08',
        "Сентября": '09',
        "Октября": '10',
        "Ноября": '11',
        "Декабря": '12',
    }
    date_month = date_dict[date[1]]
    date[1] = date_month
    date = '.'.join(date)
    return date

def convert_xls_to_xlsx(xls_file_path):
    '''Перезаписывает накладную из xls в xlsx'''
    book_xls = xlrd.open_workbook(xls_file_path, encoding_override='Windows-1251')
    book_xlsx = openpyxl.Workbook()
    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)
        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

    dst_file_path = xls_file_path + 'x'
    book_xlsx.save(dst_file_path)

def xlsx_to_json(file_path):
    df = pd.read_excel(file_path, engine='openpyxl')
    json_data = df.to_json()
    return json.loads(json_data)


def correct_names_columns(data, num_correct):
    """Корректирует словарь, приводя его к именам и значениям оригинальной таблицы"""
    dict_alfavit = {
                    0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F', 6: 'G', 7: 'H', 8: 'I', 9: 'J',
                    10: 'K', 11: 'L', 12: 'M', 13: 'N', 14: 'O', 15: 'P', 16: 'Q', 17: 'R',
                    18: 'S', 19: 'T', 20: 'U', 21: 'V', 22: 'W', 23: 'X', 24: 'Y', 25: 'Z',
                    }
    correct_dict = {}

    """Заменяем значения 'Unnamed, соответствующим буквам алфавита"""
    for key in data:
        int_key = int(str(key).split()[-1])
        for k in dict_alfavit:
            if int(k) == int_key:
                correct_dict[dict_alfavit[k]] = data[key]

    """Прописываем корректные кординаты ячеек(соответствующие документу из спам (+2))"""
    correct_dict_2 = {}
    for key in correct_dict:
        new_string_dict = {}
        for int_key in correct_dict[key]:
            # увеличиваети значение ключа (№строки) на num_correct (2 для спам, 1 для инбокс)
            new_string_dict[int(int_key) + num_correct] = correct_dict[key][int_key]
        correct_dict_2[key] = new_string_dict

    correct_dict = correct_dict_2
    return correct_dict

def post_json_request(request):
    headers = {'Content-type': 'application/json',
               'Accept': 'text/plain',
               'Content-Encoding': 'utf-8'}
    response = requests.post('https://серверполучениязапросов.рф', json=request, )#headers=headers
    print(response)