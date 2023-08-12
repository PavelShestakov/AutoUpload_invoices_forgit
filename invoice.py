import os
import xlrd
import config
import openpyxl

def convert_xls_to_xlsx(xls_file_path): #xls_file_path = путь до накладной
    '''Перезаписывает накладную из xls в xlsx, координаты данных совпадают, но накладная не копируется'''
    book_xls = xlrd.open_workbook(xls_file_path, encoding_override='Windows-1251')
    book_xlsx = openpyxl.Workbook()
    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)
        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    dst_file_path = xls_file_path + 'x'
    book_xlsx.save(dst_file_path)

def invoices_lists(dictpath=config.inbox_dict_path):
    """Функция возвращает список всех накладных и xlsx накладных в папке(аргумент)"""
    invoices_list = []
    xlsx_files = []
    all_files = os.listdir(dictpath)
    for file_name in all_files:
        file_name = config.inbox_dict_path + '\\' + file_name
        invoices_list.append(file_name)
        if file_name[-1] == 'x':
            xlsx_files.append(file_name)
    return invoices_list, xlsx_files

class Invoice():
    def __init__(self, path_to_invoice):
        self.file_name = path_to_invoice

        self.book = openpyxl.load_workbook(filename=self.file_name)

        self.sheet = [sheet_name for sheet_name in self.book][0]

        self.number = self.sheet['H14'].value
        self.date = self.sheet['J14'].value
        self.seller = self.sheet['B8'].value
        self.buyer = self.sheet['D7'].value
